"""
admin/laporan.py - Panel Rekap Laporan Pembelian (Redesign Premium)
Aplikasi Business Center SMKN 13 Bandung
Palet: Dark Green #051F20 → Pale Mint #DAF1DE
"""

import sys
import os
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from datetime import datetime, timedelta
from db import get_db

# ─── Palet Warna (seragam seluruh aplikasi) ───────────────────
C_DARKEST   = "#051F20"
C_DARK      = "#0B2B26"
C_MID       = "#163832"
C_MUTED     = "#235347"
C_MINT      = "#8EB69B"
C_PALE      = "#DAF1DE"
WHITE       = "#FFFFFF"
BG_MAIN     = "#F4FAF6"
BORDER_CLR  = "#D1E8D8"
DARK_TEXT   = "#1A2E22"
GRAY_TEXT   = "#5C7A68"
LIGHT_TEXT  = "#9DB8A8"

# Status colors
S_PENDING   = "#C07A00"   # amber
S_DITERIMA  = "#163832"   # hijau gelap
S_DITOLAK   = "#8B1A1A"   # merah gelap

S_BG_PENDING  = "#FFF8E1"
S_BG_DITERIMA = "#F0FAF4"
S_BG_DITOLAK  = "#FFF0F0"

# Card colors for Stats
STAT_BG_BLUE  = "#0B2B26"
STAT_BG_GREEN = "#163832"
STAT_BG_AMBER = "#235347"


# ─── Helper: pill button ──────────────────────────────────────────────────────
def _pill(parent, text, command, w=150, h=38, r=19,
          color=C_MID, hover=C_DARK, fg=WHITE,
          font=("Segoe UI", 9, "bold")):
    cv = tk.Canvas(parent, width=w, height=h,
                   bg=parent["bg"], highlightthickness=0, cursor="hand2")
    def _draw(fill):
        cv.delete("all")
        cv.create_arc(0,0,r*2,h,start=90,extent=180,fill=fill,outline=fill)
        cv.create_arc(w-r*2,0,w,h,start=270,extent=180,fill=fill,outline=fill)
        cv.create_rectangle(r,0,w-r,h,fill=fill,outline=fill)
        cv.create_text(w//2,h//2,text=text,fill=fg,font=font,anchor="center")
    _draw(color)
    cv.bind("<Enter>",    lambda e: _draw(hover))
    cv.bind("<Leave>",    lambda e: _draw(color))
    cv.bind("<Button-1>", lambda e: command())
    return cv


class LaporanPanel(tk.Frame):
    def __init__(self, parent, dashboard):
        super().__init__(parent, bg=BG_MAIN)
        self.dashboard   = dashboard
        self._filter_val = "1_bulan" # default: 1 bulan
        self.filtered_orders = []
        self._build_styles()
        self._build()
        self._load_laporan()

    def _build_styles(self):
        s = ttk.Style()
        s.theme_use("default")

        # Treeview pesanan
        s.configure("Laporan.Treeview",
                    font=("Segoe UI", 9), rowheight=32,
                    background=WHITE, fieldbackground=WHITE,
                    foreground=DARK_TEXT, borderwidth=0)
        s.configure("Laporan.Treeview.Heading",
                    font=("Segoe UI", 9, "bold"),
                    background=BG_MAIN, foreground=GRAY_TEXT,
                    relief="flat", borderwidth=0)
        s.map("Laporan.Treeview",
              background=[("selected", C_PALE)],
              foreground=[("selected", C_DARKEST)])

        # Scrollbar tipis
        s.configure("Thin.Vertical.TScrollbar",
                    gripcount=0, background=C_MINT,
                    troughcolor=BG_MAIN, borderwidth=0,
                    arrowsize=0, width=5)
        s.map("Thin.Vertical.TScrollbar",
              background=[("active", C_MUTED)])

    def _build(self):
        self._build_topbar()
        self._build_filter_tabs()
        tk.Frame(self, bg=BORDER_CLR, height=1).pack(fill="x")
        self._build_stats_section()
        self._build_body()

    def _build_topbar(self):
        topbar = tk.Frame(self, bg=WHITE, height=64)
        topbar.pack(fill="x")
        topbar.pack_propagate(False)

        tk.Frame(topbar, bg=C_MID, width=4).pack(side="left", fill="y")

        title_col = tk.Frame(topbar, bg=WHITE)
        title_col.pack(side="left", padx=20, fill="y", pady=12)
        tk.Label(title_col, text="Rekap Laporan Pembelian",
                 font=("Segoe UI", 14, "bold"),
                 bg=WHITE, fg=DARK_TEXT, anchor="w").pack(anchor="w")
        tk.Label(title_col, text="Unduh rekap laporan transaksi ke PDF dan Excel",
                 font=("Segoe UI", 8),
                 bg=WHITE, fg=GRAY_TEXT, anchor="w").pack(anchor="w")

        btn_ref = _pill(topbar, text="↻  Refresh", command=self._refresh,
                        w=100, h=32, r=16,
                        color=C_PALE, hover=C_MINT, fg=C_MID,
                        font=("Segoe UI", 8, "bold"))
        btn_ref.config(bg=WHITE)
        btn_ref.pack(side="right", padx=20, pady=16)

    def _build_filter_tabs(self):
        tab_wrap = tk.Frame(self, bg=WHITE)
        tab_wrap.pack(fill="x", padx=20, pady=(10,0))

        self.filter_btns = {}
        filters = [
            ("📅  1 Bulan Terakhir", "1_bulan"),
            ("📅  3 Bulan Terakhir", "3_bulan"),
            ("📅  1 Tahun Terakhir", "1_tahun"),
            ("📊  Semua Histori",   "semua"),
        ]
        for label, key in filters:
            btn = tk.Label(tab_wrap, text=label,
                           font=("Segoe UI", 9, "bold"),
                           bg=WHITE, fg=LIGHT_TEXT,
                           padx=16, pady=8, cursor="hand2")
            btn.pack(side="left")
            btn.bind("<Button-1>", lambda e, k=key: self._set_filter(k))
            self.filter_btns[key] = btn

        self._set_filter("1_bulan", init=True)

    def _build_stats_section(self):
        self.stats_frame = tk.Frame(self, bg=BG_MAIN)
        self.stats_frame.pack(fill="x", padx=20, pady=(14,6))

        # We will create 3 cards
        self.stat_cards = []
        card_labels = [
            ("Total Pendapatan", "Rp 0", STAT_BG_GREEN, "💸"),
            ("Total Transaksi Sukses", "0 Transaksi", STAT_BG_BLUE, "🛍️"),
            ("Total Item Terjual", "0 Item", STAT_BG_AMBER, "📦")
        ]

        for idx, (label, val, bg, icon) in enumerate(card_labels):
            card = tk.Frame(self.stats_frame, bg=bg, height=86,
                            highlightbackground=BORDER_CLR, highlightthickness=1)
            card.grid(row=0, column=idx, padx=(0 if idx==0 else 12, 0), sticky="ew")
            card.grid_propagate(False)
            card.pack_propagate(False)

            inner = tk.Frame(card, bg=bg)
            inner.pack(expand=True, fill="both", padx=16, pady=10)

            # Left block: labels
            lbl_col = tk.Frame(inner, bg=bg)
            lbl_col.pack(side="left", fill="both", expand=True)

            l_lbl = tk.Label(lbl_col, text=label, font=("Segoe UI", 8), bg=bg, fg=C_PALE)
            l_lbl.pack(anchor="w")

            v_lbl = tk.Label(lbl_col, text=val, font=("Segoe UI", 16, "bold"), bg=bg, fg=WHITE)
            v_lbl.pack(anchor="w", pady=(2, 0))

            # Right block: Icon
            r_lbl = tk.Label(inner, text=icon, font=("Segoe UI Emoji", 20), bg=bg, fg=WHITE)
            r_lbl.pack(side="right", padx=(8, 0))

            self.stat_cards.append((v_lbl, l_lbl))
            self.stats_frame.grid_columnconfigure(idx, weight=1, uniform="card")

    def _build_body(self):
        body = tk.Frame(self, bg=BG_MAIN)
        body.pack(fill="both", expand=True, padx=20, pady=(6,16))

        # Panel Kiri: Tabel Transaksi
        left = tk.Frame(body, bg=WHITE, highlightbackground=BORDER_CLR, highlightthickness=1)
        left.pack(side="left", fill="both", expand=True, padx=(0,12))

        l_hdr = tk.Frame(left, bg=WHITE)
        l_hdr.pack(fill="x", padx=14, pady=(12,8))
        tk.Frame(l_hdr, bg=C_MID, width=4, height=16).pack(side="left")
        tk.Label(l_hdr, text="  Detail Transaksi", font=("Segoe UI", 10, "bold"), bg=WHITE, fg=DARK_TEXT).pack(side="left")
        
        self.lbl_tabel_count = tk.Label(l_hdr, text="", font=("Segoe UI", 8), bg=WHITE, fg=GRAY_TEXT)
        self.lbl_tabel_count.pack(side="left", padx=6)

        tk.Frame(left, bg=BORDER_CLR, height=1).pack(fill="x")

        # Treeview Laporan
        tree_wrap = tk.Frame(left, bg=WHITE)
        tree_wrap.pack(fill="both", expand=True)

        cols = ("ID", "Tanggal", "Pembeli", "Item", "Metode", "Total")
        self.tree = ttk.Treeview(tree_wrap, columns=cols, show="headings", style="Laporan.Treeview")
        self.tree.heading("ID",       text="ID Pesanan")
        self.tree.heading("Tanggal",  text="Tanggal & Jam")
        self.tree.heading("Pembeli",  text="Pembeli")
        self.tree.heading("Item",     text="Item (Pcs)")
        self.tree.heading("Metode",   text="Metode")
        self.tree.heading("Total",    text="Total Belanja")

        self.tree.column("ID",      width=60,  anchor="center", stretch=False)
        self.tree.column("Tanggal", width=135, anchor="center")
        self.tree.column("Pembeli", width=110, anchor="w")
        self.tree.column("Item",    width=140, anchor="w")
        self.tree.column("Metode",  width=80,  anchor="center", stretch=False)
        self.tree.column("Total",   width=100,  anchor="e", stretch=False)

        tsb = ttk.Scrollbar(tree_wrap, orient="vertical", command=self.tree.yview, style="Thin.Vertical.TScrollbar")
        self.tree.configure(yscrollcommand=tsb.set)
        tsb.pack(side="right", fill="y", pady=4)
        self.tree.pack(fill="both", expand=True, padx=(8,0), pady=4)

        # Panel Kanan: Menu Export & Ringkasan
        right = tk.Frame(body, bg=WHITE, width=280, highlightbackground=BORDER_CLR, highlightthickness=1)
        right.pack(side="right", fill="y")
        right.pack_propagate(False)

        # Header detail rekap
        d_hdr = tk.Frame(right, bg=C_DARKEST, height=48)
        d_hdr.pack(fill="x")
        d_hdr.pack_propagate(False)
        tk.Label(d_hdr, text="Export & Cetak Laporan", font=("Segoe UI", 10, "bold"), bg=C_DARKEST, fg=WHITE).pack(side="left", padx=16, pady=13)

        info = tk.Frame(right, bg=WHITE)
        info.pack(fill="x", padx=16, pady=(16,8))

        self.lbl_info_periode = tk.Label(info, text="Periode: -", font=("Segoe UI", 10, "bold"), bg=WHITE, fg=DARK_TEXT, anchor="w")
        self.lbl_info_periode.pack(anchor="w")

        self.lbl_info_range = tk.Label(info, text="", font=("Segoe UI", 8), bg=WHITE, fg=GRAY_TEXT, anchor="w")
        self.lbl_info_range.pack(anchor="w", pady=(2, 0))

        tk.Frame(right, bg=BORDER_CLR, height=1).pack(fill="x", padx=16, pady=10)

        # Download Buttons
        btn_wrap = tk.Frame(right, bg=WHITE)
        btn_wrap.pack(fill="x", padx=16, pady=10)

        tk.Label(btn_wrap, text="Format Laporan:", font=("Segoe UI", 8, "bold"), bg=WHITE, fg=GRAY_TEXT).pack(anchor="w", pady=(0, 6))

        # PDF Button
        b_pdf = _pill(btn_wrap, text="📄  Unduh Laporan PDF", command=self._export_pdf,
                      w=246, h=44, r=22, color=C_MID, hover=C_DARK, fg=WHITE,
                      font=("Segoe UI", 9, "bold"))
        b_pdf.config(bg=WHITE)
        b_pdf.pack(pady=(0, 10))

        # Excel Button
        b_excel = _pill(btn_wrap, text="🟢  Unduh Laporan Excel", command=self._export_excel,
                        w=246, h=44, r=22, color="#2E7D32", hover="#1B5E20", fg=WHITE,
                        font=("Segoe UI", 9, "bold"))
        b_excel.config(bg=WHITE)
        b_excel.pack()

        # Instructions / Help
        help_wrap = tk.Frame(right, bg="#F0FAF4", highlightbackground=BORDER_CLR, highlightthickness=1)
        help_wrap.pack(fill="both", expand=True, padx=16, pady=(20, 16))
        
        help_inner = tk.Frame(help_wrap, bg="#F0FAF4")
        help_inner.pack(fill="both", expand=True, padx=12, pady=12)

        tk.Label(help_inner, text="💡 Informasi Laporan", font=("Segoe UI", 9, "bold"), bg="#F0FAF4", fg=C_MID, anchor="w").pack(anchor="w")
        
        instruction_text = (
            "Laporan rekap pembelian ini secara otomatis menyaring transaksi dengan "
            "status pembayaran lunas (PAID).\n\n"
            "Gunakan tombol di atas untuk mengunduh rekap dalam format PDF resmi "
            "untuk dicetak atau Excel (.xlsx) untuk pengolahan data lebih lanjut."
        )
        tk.Label(help_inner, text=instruction_text, font=("Segoe UI", 8), bg="#F0FAF4", fg=GRAY_TEXT,
                 justify="left", wraplength=218).pack(anchor="w", pady=(6,0))

    def _set_filter(self, key: str, init: bool = False):
        self._filter_val = key
        for k, btn in self.filter_btns.items():
            if k == key:
                btn.config(fg=C_MID, font=("Segoe UI", 9, "bold"))
            else:
                btn.config(fg=LIGHT_TEXT, font=("Segoe UI", 9))
        if not init:
            self._load_laporan()

    def _load_laporan(self):
        for row in self.tree.get_children():
            self.tree.delete(row)

        try:
            db = get_db()
            # Ambil semua data pesanan
            docs = db.collection('pesanan').get()

            now = datetime.now()
            start_date = None

            if self._filter_val == "1_bulan":
                start_date = now - timedelta(days=30)
                self.lbl_info_periode.config(text="Periode: 1 Bulan Terakhir")
                self.lbl_info_range.config(text=f"{start_date.strftime('%d/%m/%Y')} - {now.strftime('%d/%m/%Y')}")
            elif self._filter_val == "3_bulan":
                start_date = now - timedelta(days=90)
                self.lbl_info_periode.config(text="Periode: 3 Bulan Terakhir")
                self.lbl_info_range.config(text=f"{start_date.strftime('%d/%m/%Y')} - {now.strftime('%d/%m/%Y')}")
            elif self._filter_val == "1_tahun":
                start_date = now - timedelta(days=365)
                self.lbl_info_periode.config(text="Periode: 1 Tahun Terakhir")
                self.lbl_info_range.config(text=f"{start_date.strftime('%d/%m/%Y')} - {now.strftime('%d/%m/%Y')}")
            else:
                self.lbl_info_periode.config(text="Periode: Semua Histori")
                self.lbl_info_range.config(text="Dari awal pembukuan")

            rows = []
            total_pendapatan = 0.0
            total_transaksi = 0
            total_items = 0

            for doc in docs:
                r = doc.to_dict()
                r["id_pesanan"] = doc.id
                
                # Check payment status - only count paid orders
                pay_st = r.get("payment_status", "unpaid").lower()
                status = r.get("status", "").lower()
                
                # Report only paid and accepted orders
                if pay_st != "paid" and status != "diterima":
                    continue

                tgl_str = r.get("tanggal", "")
                try:
                    tgl_dt = datetime.strptime(tgl_str[:19], "%Y-%m-%d %H:%M:%S")
                except Exception:
                    try:
                        tgl_dt = datetime.strptime(tgl_str[:10], "%Y-%m-%d")
                    except Exception:
                        continue

                # Filter by date
                if start_date and tgl_dt < start_date:
                    continue

                rows.append((r, tgl_dt))
                total_pendapatan += float(r.get("total_harga", 0))
                total_transaksi += 1
                
                # Count details
                details = r.get("detail_pesanan", [])
                for d in details:
                    total_items += int(d.get("jumlah", 0))

            # Sort by date descending
            rows.sort(key=lambda x: x[1], reverse=True)
            self.filtered_orders = [x[0] for x in rows]

            # Update stats
            self.stat_cards[0][0].config(text=f"Rp {total_pendapatan:,.0f}")
            self.stat_cards[1][0].config(text=f"{total_transaksi} Trx")
            self.stat_cards[2][0].config(text=f"{total_items} Pcs")

            self.lbl_tabel_count.config(text=f"({len(self.filtered_orders)} transaksi)")

            # Populate Treeview
            for r in self.filtered_orders:
                total_val = f"Rp {r.get('total_harga', 0):,.0f}"
                tgl_val   = str(r.get("tanggal", ""))[:19]
                pembeli   = r.get("nama_pembeli", "-")
                metode    = (r.get("payment_method") or "-").upper()
                
                # Format detail item: e.g. "Roti (2), Susu (1)"
                details = r.get("detail_pesanan", [])
                item_parts = []
                for d in details:
                    item_parts.append(f"{d.get('nama_barang', '')} ({d.get('jumlah', 0)})")
                item_summary = ", ".join(item_parts)
                if len(item_summary) > 35:
                    item_summary = item_summary[:32] + "..."

                self.tree.insert("", "end", iid=r["id_pesanan"],
                                 values=(
                                     r["id_pesanan"][:8],
                                     tgl_val,
                                     pembeli,
                                     item_summary,
                                     metode,
                                     total_val
                                 ))

        except Exception as e:
            messagebox.showerror("Error DB", f"Gagal memuat rekap laporan: {e}", parent=self)

    def _export_pdf(self):
        if not self.filtered_orders:
            messagebox.showwarning("Peringatan", "Tidak ada data transaksi untuk diekspor pada periode ini.", parent=self)
            return

        file_path = filedialog.asksaveasfilename(
            parent=self,
            title="Simpan Laporan PDF",
            defaultextension=".pdf",
            filetypes=[("PDF Documents", "*.pdf")],
            initialfile=f"rekap_laporan_pembelian_{self._filter_val}_{datetime.now().strftime('%Y%m%d')}.pdf"
        )
        if not file_path:
            return

        try:
            from reportlab.lib.pagesizes import letter
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib import colors

            # Set up page template
            doc = SimpleDocTemplate(
                file_path,
                pagesize=letter,
                rightMargin=36, leftMargin=36,
                topMargin=36, bottomMargin=36
            )
            
            story = []
            styles = getSampleStyleSheet()

            # Styles Custom
            style_head = ParagraphStyle(
                'DocHead',
                parent=styles['Normal'],
                fontName='Helvetica-Bold',
                fontSize=14,
                leading=18,
                textColor=colors.HexColor("#051F20"),
                alignment=1 # Center
            )
            
            style_subhead = ParagraphStyle(
                'DocSubHead',
                parent=styles['Normal'],
                fontName='Helvetica',
                fontSize=9,
                leading=12,
                textColor=colors.HexColor("#5C7A68"),
                alignment=1 # Center
            )

            style_section = ParagraphStyle(
                'DocSec',
                parent=styles['Normal'],
                fontName='Helvetica-Bold',
                fontSize=11,
                leading=14,
                textColor=colors.HexColor("#163832"),
                spaceAfter=6
            )

            style_cell = ParagraphStyle(
                'CellNormal',
                parent=styles['Normal'],
                fontName='Helvetica',
                fontSize=8,
                leading=10,
                textColor=colors.HexColor("#1A2E22")
            )

            style_cell_bold = ParagraphStyle(
                'CellBold',
                parent=styles['Normal'],
                fontName='Helvetica-Bold',
                fontSize=8,
                leading=10,
                textColor=colors.HexColor("#051F20")
            )

            # 1. Header Yayasan / Sekolah
            story.append(Paragraph("BUSINESS CENTER SMKN 13 BANDUNG", style_head))
            story.append(Paragraph("Sistem Manajemen Kasir Modern  •  Jl. Soekarno-Hatta Km. 10, Bandung", style_subhead))
            story.append(Spacer(1, 10))
            story.append(HRFlowable(width="100%", thickness=1.5, color=colors.HexColor("#8EB69B"), spaceAfter=14))

            # 2. Judul Laporan
            story.append(Paragraph("LAPORAN REKAP PEMBELIAN / PENJUALAN", style_section))
            
            # Info Laporan Block
            tgl_cetak = datetime.now().strftime("%d %B %Y %H:%M:%S")
            periode_str = {
                "1_bulan": "1 Bulan Terakhir",
                "3_bulan": "3 Bulan Terakhir",
                "1_tahun": "1 Tahun Terakhir",
                "semua": "Semua Periode"
            }.get(self._filter_val, self._filter_val)

            total_sales = sum(float(x.get("total_harga", 0)) for x in self.filtered_orders)
            total_items = 0
            for x in self.filtered_orders:
                for d in x.get("detail_pesanan", []):
                    total_items += int(d.get("jumlah", 0))

            info_data = [
                [Paragraph(f"<b>Periode:</b> {periode_str}", style_cell), Paragraph(f"<b>Tanggal Cetak:</b> {tgl_cetak}", style_cell)],
                [Paragraph(f"<b>Total Pendapatan:</b> Rp {total_sales:,.0f}", style_cell), Paragraph(f"<b>Total Item Terjual:</b> {total_items} Pcs", style_cell)],
                [Paragraph(f"<b>Total Transaksi:</b> {len(self.filtered_orders)} Transaksi Sukses", style_cell), ""]
            ]
            info_table = Table(info_data, colWidths=[270, 270])
            info_table.setStyle(TableStyle([
                ('VALIGN', (0,0), (-1,-1), 'TOP'),
                ('BOTTOMPADDING', (0,0), (-1,-1), 4),
                ('TOPPADDING', (0,0), (-1,-1), 0),
            ]))
            story.append(info_table)
            story.append(Spacer(1, 16))

            # 3. Tabel Detail
            # Headers
            table_headers = [
                Paragraph("<b>No</b>", style_cell_bold),
                Paragraph("<b>ID Order</b>", style_cell_bold),
                Paragraph("<b>Tanggal & Jam</b>", style_cell_bold),
                Paragraph("<b>Nama Pembeli</b>", style_cell_bold),
                Paragraph("<b>Detail Item (Qty)</b>", style_cell_bold),
                Paragraph("<b>Metode</b>", style_cell_bold),
                Paragraph("<b>Total Harga</b>", style_cell_bold)
            ]
            
            table_rows = [table_headers]
            
            for idx, r in enumerate(self.filtered_orders, 1):
                details = r.get("detail_pesanan", [])
                item_parts = []
                for d in details:
                    item_parts.append(f"{d.get('nama_barang', '')} ({d.get('jumlah', 0)})")
                items_str = ", ".join(item_parts)

                table_rows.append([
                    Paragraph(str(idx), style_cell),
                    Paragraph(r["id_pesanan"][:8], style_cell),
                    Paragraph(str(r.get("tanggal", ""))[:19], style_cell),
                    Paragraph(r.get("nama_pembeli", "-"), style_cell),
                    Paragraph(items_str, style_cell),
                    Paragraph(str(r.get("payment_method") or "-").upper(), style_cell),
                    Paragraph(f"Rp {r.get('total_harga', 0):,.0f}", style_cell)
                ])

            # Row for Total
            table_rows.append([
                Paragraph("<b>TOTAL KESELURUHAN</b>", style_cell_bold), "", "", "", "", "",
                Paragraph(f"<b>Rp {total_sales:,.0f}</b>", style_cell_bold)
            ])

            # Widths: page is 612 wide. Margins 36+36 = 72. Printable width = 540.
            col_widths = [24, 52, 100, 90, 164, 50, 60]
            
            main_table = Table(table_rows, colWidths=col_widths, repeatRows=1)
            
            # Table styles
            t_style = TableStyle([
                # Header formatting
                ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#DAF1DE")),
                ('ALIGN', (0,0), (-1,-1), 'LEFT'),
                ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
                ('BOTTOMPADDING', (0,0), (-1,-1), 6),
                ('TOPPADDING', (0,0), (-1,-1), 6),
                ('GRID', (0,0), (-1,-2), 0.5, colors.HexColor("#D1E8D8")),
                
                # Grand Total row formatting
                ('SPAN', (0, -1), (5, -1)),
                ('ALIGN', (0, -1), (5, -1), 'RIGHT'),
                ('BACKGROUND', (0,-1), (-1,-1), colors.HexColor("#EEF8F1")),
                ('LINEABOVE', (0,-1), (-1,-1), 1.5, colors.HexColor("#8EB69B")),
                ('LINEBELOW', (0,-1), (-1,-1), 1.5, colors.HexColor("#8EB69B")),
            ])
            
            main_table.setStyle(t_style)
            story.append(main_table)
            story.append(Spacer(1, 40))

            # 4. Signature Block
            sig_data = [
                ["", "Bandung, " + datetime.now().strftime("%d %B %Y")],
                ["", "Mengetahui,"],
                ["", "Kepala Business Center SMKN 13"],
                ["", Spacer(1, 45)],
                ["", "______________________________"],
                ["", "NIP. 19780512 200501 2 003"]
            ]
            sig_table = Table(sig_data, colWidths=[300, 240])
            sig_table.setStyle(TableStyle([
                ('ALIGN', (1,0), (1,-1), 'CENTER'),
                ('VALIGN', (0,0), (-1,-1), 'TOP'),
                ('FONTNAME', (1,0), (1,-1), 'Helvetica'),
                ('FONTSIZE', (1,0), (1,-1), 9),
            ]))
            story.append(sig_table)

            # Build Document
            doc.build(story)
            messagebox.showinfo("Sukses Ekspor", f"Laporan rekap berhasil diunduh ke:\n{file_path}", parent=self)

        except Exception as e:
            messagebox.showerror("Error Ekspor PDF", f"Gagal mengekspor PDF: {e}", parent=self)

    def _export_excel(self):
        if not self.filtered_orders:
            messagebox.showwarning("Peringatan", "Tidak ada data transaksi untuk diekspor pada periode ini.", parent=self)
            return

        file_path = filedialog.asksaveasfilename(
            parent=self,
            title="Simpan Laporan Excel",
            defaultextension=".xlsx",
            filetypes=[("Excel Files", "*.xlsx")],
            initialfile=f"rekap_laporan_pembelian_{self._filter_val}_{datetime.now().strftime('%Y%m%d')}.xlsx"
        )
        if not file_path:
            return

        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Rekap Pembelian"
            
            # Show grid lines
            ws.views.sheetView[0].showGridLines = True

            # Colors & Styles
            COLOR_DARK_GREEN = "163832"
            COLOR_LIGHT_GREEN = "DAF1DE"
            COLOR_ZEBRA = "F4FAF6"

            font_title = Font(name="Segoe UI", size=16, bold=True, color="051F20")
            font_subtitle = Font(name="Segoe UI", size=10, italic=True, color="5C7A68")
            font_section = Font(name="Segoe UI", size=11, bold=True, color="163832")
            font_header = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
            font_body = Font(name="Segoe UI", size=10)
            font_bold = Font(name="Segoe UI", size=10, bold=True)
            
            fill_header = PatternFill(start_color=COLOR_DARK_GREEN, end_color=COLOR_DARK_GREEN, fill_type="solid")
            fill_zebra = PatternFill(start_color=COLOR_ZEBRA, end_color=COLOR_ZEBRA, fill_type="solid")
            fill_total = PatternFill(start_color="EEF8F1", end_color="EEF8F1", fill_type="solid")
            
            border_thin = Side(border_style="thin", color="D1E8D8")
            border_double = Side(border_style="double", color="8EB69B")
            border_thick = Side(border_style="medium", color="163832")
            
            border_all = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)
            border_total_row = Border(top=border_thin, bottom=border_double)

            # 1. Header Title
            ws["A1"] = "BUSINESS CENTER SMKN 13 BANDUNG"
            ws["A1"].font = font_title
            
            ws["A2"] = "Laporan Rekap Pembelian Kasir (Lunas / Sukses)"
            ws["A2"].font = font_subtitle
            ws.row_dimensions[1].height = 24
            ws.row_dimensions[2].height = 18

            # 2. Metadata Info
            periode_str = {
                "1_bulan": "1 Bulan Terakhir",
                "3_bulan": "3 Bulan Terakhir",
                "1_tahun": "1 Tahun Terakhir",
                "semua": "Semua Periode"
            }.get(self._filter_val, self._filter_val)

            ws["A4"] = "Periode:"
            ws["B4"] = periode_str
            ws["A5"] = "Dibuat Pada:"
            ws["B5"] = datetime.now().strftime("%d %B %Y %H:%M:%S")

            ws["A4"].font = font_bold
            ws["B4"].font = font_body
            ws["A5"].font = font_bold
            ws["B5"].font = font_body

            # Statistics block
            total_sales = sum(float(x.get("total_harga", 0)) for x in self.filtered_orders)
            total_items = 0
            for x in self.filtered_orders:
                for d in x.get("detail_pesanan", []):
                    total_items += int(d.get("jumlah", 0))

            ws["D4"] = "Total Transaksi:"
            ws["E4"] = len(self.filtered_orders)
            ws["D5"] = "Total Produk Terjual:"
            ws["E5"] = total_items
            ws["D6"] = "Total Pendapatan:"
            ws["E6"] = total_sales
            ws["E6"].number_format = '"Rp"#,##0'

            for row in range(4, 7):
                ws[f"D{row}"].font = font_bold
                ws[f"E{row}"].font = font_body
            ws["E6"].font = font_bold

            # 3. Table Headers
            headers = ["No", "ID Pesanan", "Tanggal & Waktu", "Nama Pembeli", "Metode Bayar", "Detail Item Dipesan (Qty)", "Total Belanja"]
            start_row = 8
            
            for col_idx, text in enumerate(headers, 1):
                cell = ws.cell(row=start_row, column=col_idx, value=text)
                cell.font = font_header
                cell.fill = fill_header
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = Border(left=border_thin, right=border_thin, top=border_thick, bottom=border_thick)
            
            ws.row_dimensions[start_row].height = 28

            # 4. Table Body
            current_row = start_row + 1
            for idx, r in enumerate(self.filtered_orders, 1):
                details = r.get("detail_pesanan", [])
                item_parts = []
                for d in details:
                    item_parts.append(f"{d.get('nama_barang', '')} ({d.get('jumlah', 0)})")
                items_str = ", ".join(item_parts)

                ws.cell(row=current_row, column=1, value=idx)
                ws.cell(row=current_row, column=2, value=r["id_pesanan"][:8])
                ws.cell(row=current_row, column=3, value=str(r.get("tanggal", ""))[:19])
                ws.cell(row=current_row, column=4, value=r.get("nama_pembeli", "-"))
                ws.cell(row=current_row, column=5, value=str(r.get("payment_method") or "-").upper())
                ws.cell(row=current_row, column=6, value=items_str)
                total_cell = ws.cell(row=current_row, column=7, value=float(r.get("total_harga", 0)))
                
                # Format
                total_cell.number_format = '"Rp"#,##0'
                
                # Alignments
                ws.cell(row=current_row, column=1).alignment = Alignment(horizontal="center")
                ws.cell(row=current_row, column=2).alignment = Alignment(horizontal="center")
                ws.cell(row=current_row, column=3).alignment = Alignment(horizontal="center")
                ws.cell(row=current_row, column=5).alignment = Alignment(horizontal="center")
                
                # Border & Zebra
                for col_idx in range(1, 8):
                    cell = ws.cell(row=current_row, column=col_idx)
                    cell.font = font_body
                    cell.border = border_all
                    if idx % 2 == 0:
                        cell.fill = fill_zebra
                
                ws.row_dimensions[current_row].height = 22
                current_row += 1

            # 5. Grand Total Row
            ws.cell(row=current_row, column=1, value="TOTAL KESELURUHAN")
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)
            ws.cell(row=current_row, column=1).alignment = Alignment(horizontal="right", vertical="center")
            
            total_sum_cell = ws.cell(row=current_row, column=7, value=f"=SUM(G{start_row+1}:G{current_row-1})")
            total_sum_cell.number_format = '"Rp"#,##0'
            
            ws.row_dimensions[current_row].height = 24
            
            # Format total row
            for col_idx in range(1, 8):
                cell = ws.cell(row=current_row, column=col_idx)
                cell.font = font_bold
                cell.fill = fill_total
                cell.border = border_total_row

            # 6. Column Widths Auto-fit
            for col in ws.columns:
                max_len = 0
                col_letter = openpyxl.utils.get_column_letter(col[0].column)
                
                # Don't size based on title (row 1 & 2) or merged cell in total
                for cell in col:
                    if cell.row > 2 and cell.value and cell.row != current_row:
                        val_str = str(cell.value)
                        if cell.column == 7: # Currency length approx
                            val_str = f"Rp {cell.value:,.0f}" if isinstance(cell.value, (int, float)) else str(cell.value)
                        if len(val_str) > max_len:
                            max_len = len(val_str)
                
                ws.column_dimensions[col_letter].width = max(max_len + 4, 10)

            # Special spacing for specific columns
            ws.column_dimensions['A'].width = 6
            ws.column_dimensions['B'].width = 12
            ws.column_dimensions['C'].width = 20
            ws.column_dimensions['F'].width = 36 # Detail column should be spacious

            wb.save(file_path)
            messagebox.showinfo("Sukses Ekspor", f"Laporan rekap berhasil diunduh ke:\n{file_path}", parent=self)

        except Exception as e:
            messagebox.showerror("Error Ekspor Excel", f"Gagal mengekspor Excel: {e}", parent=self)

    def _refresh(self):
        self._load_laporan()
